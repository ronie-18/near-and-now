#!/usr/bin/env python3
from openpyxl import load_workbook
import json

wb = load_workbook('PRODUCT EXCEL/Biscuits & Cookies.xlsx')
ws = wb.active

# Get headers
headers = []
for cell in ws[1]:
    if cell.value:
        headers.append(str(cell.value))

print(f"Headers: {headers}")
print("=" * 80)

# Get all rows
products = []
for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
    product = {}
    for i, value in enumerate(row):
        if i < len(headers):
            product[headers[i]] = value
    if any(product.values()):  # Only add if not empty
        products.append(product)

print(f"Total products: {len(products)}")
print("=" * 80)

# Print first few products
for i, product in enumerate(products[:3]):
    print(f"\nProduct {i+1}:")
    for key, value in product.items():
        print(f"  {key}: {value}")

# Save to JSON
with open('biscuits_cookies_data.json', 'w') as f:
    json.dump(products, f, indent=2, default=str)

print("\n✅ Data saved to biscuits_cookies_data.json")

